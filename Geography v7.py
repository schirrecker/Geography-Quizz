import xlrd
import openpyxl
import random
import datetime
import os, os.path
import pickle

# -----------------------------------------------------
# openpyxl syntax: 
# wb = openpyxl.Workbook()
# grab the active worksheet: ws = wb.active
# Data can be assigned directly to cells: ws['A1'] = 42
# Rows can also be appended: ws.append([1, 2, 3])
# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
# Save the file - wb.save("sample.xlsx")
#-----------------------------------------------------

# ----------------------------------------------------
# xlrd syntax:
# workbook = xlrd.open_workbook("FILE.xlsx")
# worksheet = workbook.sheet_by_name("<NAME OF SHEET>")
# worksheet.cell(row, col).value)
# worksheet.row_values(row)
# worksheet.nrows
# worksheet.ncols
# Row = Country, Country Code, Continent, Capital,
# Population, Area, Coastline, Government, Currency 
# ----------------------------------------------------

# ----------------------------------------------------
# Class definitions
# ----------------------------------------------------
class Player:
    def __init__(self, name, password="password"):
        self.name = name
        self.password = password
        self.points = 0
        self.level = 1
        self.knowledge = {}
        self.CountriesToTest = []
        self.InitKnowledge()
        
        # self.knowlege = {"country": [-1, -1, -1, -1, -1, -1]}
        # 0  = not asked yet
        # 1  = correct
        # -1 = wrong
         
    def InitKnowledge(self):
        for country in Countries:
            self.knowledge[country.name] = [-1, -1, -1, -1, -1, -1]

    def SaveKnowledgeData(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["Name", "Continent", "Capital", "Population", "Area", "Coastline", "Currency", "Ratio"])
        for country in Countries:
            row = [country.name]
            Nb_facts = 0
            if self.knowledge[country.name][0] == 1:
                row.append(country.continent)
                Nb_facts += 1
            else:
                row.append("still learning")
            if self.knowledge[country.name][1] == 1:
                row.append(country.capital)
                Nb_facts += 1
            else:
                row.append("still learning")
            if self.knowledge[country.name][2] == 1:
                row.append(int(country.population))
                Nb_facts += 1
            else:
                row.append("still learning")
            if self.knowledge[country.name][3] == 1:
                row.append(int(country.area))
                Nb_facts += 1
            else:
                row.append("still learning")
            if self.knowledge[country.name][4] == 1:
                row.append(int(country.coastline))
                Nb_facts += 1
            else:
                row.append("still learning")
            if self.knowledge[country.name][5] == 1:
                row.append(country.currency)
                Nb_facts += 1
            else:
                row.append("still learning")
            if Nb_facts > 0:
                row.append(str(int(100*Nb_facts/6))+"%")
                worksheet.append(row)
        workbook.save(self.name + ".xls")
        print("Your current knowledge was saved in file " + self.name + '.xls')

    def GetKnowledgeData(self):
        pass

    def PrintKnowledge(self):
        for country in Countries:
            if self.knowledge[country.name][0] == 1:    
                print (country.name + " - continent: " + country.continent)
            if self.knowledge[country.name][1] == 1:    
                print (country.name + " - capital: " + country.capital)            
            if self.knowledge[country.name][2] == 1:    
                print (country.name + " - population: " + str(int(country.population)))               
            if self.knowledge[country.name][3] == 1:    
                print (country.name + " - area: " + str(int(country.area)))              
            if self.knowledge[country.name][4] == 1:    
                print (country.name + " - coastline: " + str(int(country.coastline)))
            if self.knowledge[country.name][5] == 1:    
                print (country.name + " - currency: " + country.currency)                

    def PrintKnowledgeStats(self):
        capitals, continents, populations, areas, coastlines, currencies = 0, 0, 0, 0, 0, 0
        ratio = [0, 0, 0, 0, 0, 0, 0]
        levels = [0 for i in range(MAX_LEVEL)]
        for country in Countries:
            i = 0
            if self.knowledge[country.name][0] == 1:
                continents += 1
                i += 1
            if self.knowledge[country.name][1] == 1:
                capitals += 1
                i += 1
            if self.knowledge[country.name][2] == 1:
                populations += 1
                i += 1
            if self.knowledge[country.name][3] == 1:
                areas += 1
                i += 1
            if self.knowledge[country.name][4] == 1:
                coastlines += 1
                i += 1
            if self.knowledge[country.name][5] == 1:
                currencies += 1
                i += 1
            if i > 0:
                for lvl in range(MIN_LEVEL, MAX_LEVEL+1):
                    if country in CountriesByLevel[str(lvl)]:
                        levels[lvl-1] += 1
            ratio[i] += 1
        print()
        print ("Here is the report card for " + self.name + ":")
        print ("----------------------------" + "-" * len(self.name))
        print("You  have " + str(self.points) + " points")
        for lvl in range(MIN_LEVEL, MAX_LEVEL+1):
            print("Level " + str(lvl) + ": " + str(levels[lvl-1]) + " out of " + str(len(CountriesByLevel[str(lvl)])))
        print()
        print ("Capitals: " + str(capitals))
        print ("Continents: " + str(continents))
        print ("Population sizes: " + str(populations))
        print ("Areas: " + str(areas))
        print ("Coastlines: " + str(coastlines))
        print ("Currencies: " + str(currencies))
        print()
        for i in range(7):
            print ("Number of country with " + str(i) + " correct answers: " + str(ratio[i]))
        print()

    # test only countries that player doesn't know
    # 0: continent, 1: capital, 2: pop, 3: area, 4: coastline, 5: currency
    def UpdateCountriesToTest(self, level, continents):
        self.CountriesToTest = []
        for country in CountriesByLevel[str(level)]:
            if self.knowledge[country.name][0] != 1 and self.knowledge[country.name][1] != 1:
                if country.continent in continents:
                    self.CountriesToTest.append(country)

    def QuizRound(self, Nb_Questions, level, continents):
        self.UpdateCountriesToTest(level, continents)
        for i in range(Nb_Questions):
            country = random.choice(self.CountriesToTest)
            print ("")
            self.Quiz(country)
            print ("")
            print ("You have " + str(self.points) + " points")

    def Quiz(self, country):
        # Continent
        # ---------
        validInput = False
        while not validInput:
            try:
                txt = "In which continent is the country of " + country.name + " ? " + os.linesep
                txt = txt + "Continents are: " + " ,".join(Continents) + os.linesep
                answer = input(txt)
            except:
                print("Entry error, please try again")
            else:
                if answer in Continents:
                    validInput = True
                else:
                    print ("please enter from the list of continents")
        if country.CheckContinent(answer):
            print ("This is correct")
            self.points += 1
            self.knowledge[country.name][0] = 1
        else:
            print ("This is wrong")
            self.knowledge[country.name][0] = 0
        print (" The answer was: ", country.continent)

       # Capital
       # -------
        validInput = False
        while not validInput:
            try:
                answer = input("What is the capital of " + country.name + " ? ")
            except:
                print("Entry error, please try again")
            else:
                if len(answer) > 1 and len(answer) != answer.count(" "):
                    validInput = True
                else:
                    print ("print enter a valid city")
        if country.CheckCapital(answer):
            print ("This is correct")
            self.points += 1
            self.knowledge[country.name][1] = 1
        else:
            print ("This is wrong")
            self.knowledge[country.name][1] = 0
        print ("Capital: ", country.capital)

        # Population
        # ----------
        validInput = False
        while not validInput:
            try:
                answer = int(input("What is the population of " + country.name + " (in millions ?) "))*1000000
            except:
                print("Entry error, please enter a valid number")
            else:
                if answer > 0:
                    validInput = True
                else:
                    print("Please enter a valid number")
        if country.CheckPopulation(answer):
            print ("This is correct")
            self.points += 1
            self.knowledge[country.name][2] = 1
        else:
            print ("This is wrong")
            self.knowledge[country.name][2] = 0
        print ("Population: ", int(country.population))

        # Area
        # --------------
        validInput = False
        while not validInput:
            try:
                answer = input("What is the area of " + country.name + " ? ")
            except:
                print("Entry error, please try again")
            else:
                if len(answer) != answer.count(" "):
                    answer = int(answer)
                    validInput = True
                else:
                    print("Please enter a valid number")
        if country.CheckArea(answer):
            print ("This is correct")
            self.points += 1
            self.knowledge[country.name][3] = 1
        else:
            print ("This is wrong")
            self.knowledge[country.name][3] = 0
        print ("Area: ", int(country.area))       

        # Coastline
        # ---------
        validInput = False
        while not validInput:
            try:
                answer = input("Does " + country.name + " have a coastline ? (y or n)")
            except:
                print("Entry error, please try again")
            else:
                if answer.lower() in ["yes", "no", "y", "n"]:
                    validInput = True
                else:
                    print("Please enter a valid response")
        if country.CheckCoastline(answer):
            print ("This is correct")
            self.points += 1
            self.knowledge[country.name][4] = 1
        else:
            print ("This is wrong")
            self.knowledge[country.name][4] = 0
        print ("Coastline: ", int(country.coastline))

        # Currency
        # ---------
        validInput = False
        while not validInput:
            try:
                answer = input("What is the currency of " + country.name + " ? ")
            except:
                print("Entry error, please try again")
            else:
                if len(answer) != answer.count(" "):
                    validInput = True
                else:
                    print("Please enter a valid response")
        if country.CheckCurrency(answer):
            print ("This is correct")
            self.points += 1
            self.knowledge[country.name][5] = 1
        else:
            print ("This is wrong")
            self.knowledge[country.name][5] = 0
        print ("Currency: ", country.currency)       

class Country:
    def __init__(self, name, continent, capital, population,
                 area, coastline,currency, level):
        self.name = name
        self.continent = continent
        self.continent_margin = .9
        self.capital = capital
        self.capital_margin = .8
        self.population = population
        self.population_margin = 0.3
        self.area = area
        self.area_margin = 0.5
        self.coastline = coastline
        self.coastine_margin = .7
        self.currency = currency
        self.currency_margin = .7
        self.level = int(level)

    def __repr__(self):
        txt = self.name + os.linesep
        txt = txt + "Continent: " + self.continent + os.linesep
        txt = txt + "Capital: " + self.capital + os.linesep
        txt = txt + "Population: " + str(int(self.population/1000000)) + "M" + os.linesep
        txt = txt + "Area: " + str(self.area) + os.linesep
        txt = txt + "Coastline: " + str(self.coastline) + os.linesep
        txt = txt + "Currency: " + self.currency + os.linesep
        txt = txt + "Difficulty: " + str(self.level) + os.linesep
        return txt
      
    def CheckCapital(self, capital):
        return stringsMatch(self.capital, capital, self.capital_margin)

    def CheckContinent(self, continent):
        return stringsMatch(self.continent, continent, self.continent_margin)

    def CheckCurrency(self, currency):
        return stringsMatch(self.currency, currency, self.currency_margin)

    def CheckPopulation(self, population):
        return isclose(self.population, population, self.population_margin)
            
    def CheckArea(self, area):
        if self.area < 1000: 
            return area < 1000
        elif self.area < 10000:
            return isclose(self.area, area, 5)      
        elif self.area < 100000:
            return isclose(self.area, area, 3)     
        elif self.area < 500000:
            return isclose(self.area, area, .7)
        else:
            return isclose(self.area, area, self.area_margin)       

    def CheckCoastline(self, coastline):
        coastline = coastline.lower()[0]     # first letter y or n
        if (self.coastline == 0 and coastline == 'n') or (self.coastline != 0 and coastline == 'y'):
            return True
        else:
            return False
        
# ---------------------------------------------------
# Useful functions
# ---------------------------------------------------
def isclose(a, b, margin):
    if a * (1 - margin) < b < a * (1 + margin):
        return True
    else:
        return False

def stringsMatch(a, b, margin):
    x = a.lower()
    y = b.lower()
    hits = 0
    if x == y:
        return True
    else:
        for i in range(0, min(len(x), len(y))):
            if x[i] == y[i]:
                hits += 1
        # print ("hits: ", str(hits))
        # print ("margin * max: ", str(margin * max(len(x), len(y))))
        # l= input ("continue")
        if hits >= margin * max(len(x), len(y)):
            return True
        else:
            return False

def approx(a, n):
    return int(a/(10**n)) * (10**n)
    

# -------------------------------------------------
# Load country data from file
# -------------------------------------------------
def LoadCountryData():
    Continents = []
    Countries = []
    CountriesByLevel = {"1": [], "2": [], "3": [], "4": [], "5": []}
    for row in range(1, worksheet.nrows):
        country = worksheet.row_values(row)
        Countries.append(Country(
            country[0],               # name
            country[1],               # continent
            country[2],               # capital
            approx(country[3], 3),    # population
            country[4],               # area
            country[5],               # coastline
            country[6],               # currency
            country[8]))              # difficulty level

    # create levels: dictionary = {"level" : [list of countries]}
    # create continents: list of continents
    for country in Countries:
        CountriesByLevel[str(country.level)].append(country)
        Continents.append(country.continent)
    Continents = list(set(Continents))
    return Countries, CountriesByLevel, Continents

# -------------------------------------------------
# Save players ranking to .xls file
# -------------------------------------------------
def SaveScores():
    global Players
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Player Name", "Score"])
    for player in Players:
        worksheet.append([player.name, player.points])           
    workbook.save("scores.xls")
    print("Current scores saved in file scores.xls")

# -------------------------------------------------
# Save player data to file
# -------------------------------------------------
def SavePlayerData():
    global Players
    with open('PlayerData', 'wb') as PlayerFile:
        pickle.dump(Players, PlayerFile)

# -------------------------------------------------
# Load player data from file
# -------------------------------------------------
def LoadPlayerData():
    Players = []
    if os.path.isfile('PlayerData'):
        with open('PlayerData', 'rb') as PlayerFile:
            Players = pickle.load(PlayerFile)
            print ("Player data loaded.")
    else:
        print("No existing player data.  This is the first game.")
    return Players

# ----------------------------------------------------
# Load Excel file and its worksheets
# ----------------------------------------------------
workbook = xlrd.open_workbook("Countries.xlsx")
worksheet = workbook.sheet_by_name("Facts")
Nb_Countries = worksheet.nrows
Nb_Facts = worksheet.ncols - 1

# -------------------------------------------------
# Load data
# -------------------------------------------------
Countries, CountriesByLevel, Continents = LoadCountryData()
Players = LoadPlayerData()
NB_QUESTIONS = 1
MAX_LEVEL = 5
MIN_LEVEL = 1

print (str(worksheet.nrows) + " countries loaded.")
txt = "Existing players: "
for player in Players:
    txt = txt + player.name + " "
print (txt)

def CheckPlayer(name, Players):
    exist = False
    for p in Players:
        if name.lower() == p.name.lower():
            print ("Welcome back, " + name + os.linesep)
            exist = True
            return p
    if not exist:          # if it doesn't exist, create it
        answer = input("Would you like to create this user? (y/n) ")
        if answer[0].lower() == "y":
            p = Player(name)
            Players.append(p)
            return p

# -------------------------------------------------
# Start game loop
# -------------------------------------------------

name = str(input("Enter player name: "))
player = CheckPlayer(name, Players)
player.SaveKnowledgeData()
print ("Your are at level " + str(player.level))
print ("You currently have " + str(player.points) + " points")

quit = False 
while quit == False:
    try:
        play = input("Would you like to play ? (y/n)")
    except:
        print ("Input error, try again")
    else:
        if play.lower() == "n":
            quit = True
        else:
            level = ""
            while level not in range(MIN_LEVEL, MAX_LEVEL+1):
                level = int(input("what level? (1 to 5) "))
            test_continents = [c for c in Continents]
            test_continents.append("All")
            continent = ""
            while continent not in test_continents:
                continent = input("What continent? " + str(test_continents) + " ")
            if continent == "All":
                player.QuizRound(NB_QUESTIONS, int(level), Continents)
            else:
                player.QuizRound(NB_QUESTIONS, int(level), continent)

# -----------------------
# Quit and Save Data
# -----------------------
player.PrintKnowledgeStats()
player.SaveKnowledgeData()
SaveScores()




    
