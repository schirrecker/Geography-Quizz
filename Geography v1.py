import xlrd
import openpyxl
import random
import datetime

# -----------------------------------------------------
# openpyxl syntax: 
# wb = openpyxl.Workbook()
# grab the active worksheet: ws = wb.active
# Data can be assigned directly to cells: ws['A1'] = 42
# Rows can also be appended: ws.append([1, 2, 3])
# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
# Save the file - wb.save("sample.xlsx")
#-----------------------------------------------------

# ----------------------------------------------------
# xlrd syntax:
# workbook = xlrd.open_workbook("FILE.xlsx")
# worksheet = workbook.sheet_by_name("<NAME OF SHEET>")
# worksheet.cell(row, col).value)
# worksheet.row_values(row)
# worksheet.nrows
# worksheet.ncols
# Row = Country, Country Code, Continent, Capital,
# Population, Area, Coastline, Government, Currency 
# ----------------------------------------------------

CORRECT = 1
WRONG = 0
NOTASKED = -1
CAPITAL = 2

# ----------------------------------------------------
# Load Excel file and its worksheets
# ----------------------------------------------------
workbook = xlrd.open_workbook("Countries.xlsx")
worksheet = workbook.sheet_by_name("Facts")
Nb_Countries = worksheet.nrows
Nb_Facts = worksheet.ncols - 1

# ----------------------------------------------------
# Class definitions
# ----------------------------------------------------
class Player:
    def __init__(self, name, password="password"):
        self.name = name
        self.password = password
        self.points = 0
        self.knowledge = {}
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.BuildWorkbook()
        
        # self.knowlege = {"country": [["field1", rating],
        #                              ["field2", rating], [], ...]}
        # 0  = not asked yet
        # 1  = correct
        # -1 = wrong

    def BuildWorkbook(self):
        for row in range(1, Nb_Countries):
            newrow = []
            newrow.append(worksheet.cell(row, 0).value)   
            for i in range (1,Nb_Facts):
                newrow.append(NOTASKED)            
            self.worksheet.append(newrow)
        self.SavePlayerData()

    def SavePlayerData(self):
        self.workbook.save(self.name + ".xls")

    def GetPlayerData(self):
        pass
    
# -------------------------------------------------
# Initialize global data, define useful functions
# -------------------------------------------------
CountryData = {}
Players = []
Nb_Questions = 10

def CountryLoader():
    for row in range(worksheet.nrows):
        country = worksheet.row_values(row)
        CountryData[str(country[0])]=country[1:]
        
def AskForCapital(Player):
    country = random.choice(list(CountryData.keys()))
    answer = input("What is the capital of " + country + " ? ")
    if answer.lower() == CountryData.get(country)[2].lower():
        print ("This is correct")
        Player.points += 1
    else:
        print ("This is wrong")
        print (" The answer was: ", CountryData.get(country)[CAPITAL])
        # record "Capital" as missed item in Player's misssed list
        # Player.missed[country].append("Capital")

# -------------------------------------------------
# Load data from Excel sheet
# -------------------------------------------------
print ("Loading country data...")
CountryLoader()
print (str(worksheet.nrows) + " countries loaded.")

# -------------------------------------------------
# Select and initialize player
# -------------------------------------------------
name = str(input("Enter player name: "))
player = Player(name)

# -------------------------------------------------
# Start game loop
# -------------------------------------------------

for i in range(1, Nb_Questions):
    AskForCapital(player)


    



    
