import xlrd
import openpyxl
import random
import datetime
import os

# -----------------------------------------------------
# openpyxl syntax: 
# wb = openpyxl.Workbook()
# grab the active worksheet: ws = wb.active
# Data can be assigned directly to cells: ws['A1'] = 42
# Rows can also be appended: ws.append([1, 2, 3])
# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
# Save the file - wb.save("sample.xlsx")
#-----------------------------------------------------

# ----------------------------------------------------
# xlrd syntax:
# workbook = xlrd.open_workbook("FILE.xlsx")
# worksheet = workbook.sheet_by_name("<NAME OF SHEET>")
# worksheet.cell(row, col).value)
# worksheet.row_values(row)
# worksheet.nrows
# worksheet.ncols
# Row = Country, Country Code, Continent, Capital,
# Population, Area, Coastline, Government, Currency 
# ----------------------------------------------------


# ----------------------------------------------------
# Load Excel file and its worksheets
# ----------------------------------------------------
workbook = xlrd.open_workbook("Countries.xlsx")
worksheet = workbook.sheet_by_name("Facts")
Nb_Countries = worksheet.nrows
Nb_Facts = worksheet.ncols - 1


# ---------------------------------------------------
# Useful functions
# ---------------------------------------------------
def isclose(a, b, margin):
    if a * (1 - margin) < b < a * (1 + margin):
        return True
    else:
        return False

def stringsMatch(a, b):
    if a.lower() == b.lower():
        return True
    else:
        return False


# ----------------------------------------------------
# Class definitions
# ----------------------------------------------------
class Player:
    def __init__(self, name, password="password"):
        self.name = name
        self.password = password
        self.points = 0
        self.level = 1
        self.knowledge = {}
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.InitKnowledge()
        self.SaveData()
        
        # self.knowlege = {"country": [-1, -1, -1, -1, -1, -1]}
        # 0  = not asked yet
        # 1  = correct
        # -1 = wrong
         
    def InitKnowledge(self):
        for country in Countries:
            self.knowledge[country] = [-1, -1, -1, -1, -1, -1]

    def SaveData(self):
        for row in range(1, Nb_Countries):
            newrow = []
            country = worksheet.cell(row, 0).value
            newrow.append(country)
            newrow.extend(self.knowledge[country])
            self.worksheet.append(newrow)
        self.workbook.save(self.name + ".xls")

    def GetData(self):
        pass

class Country:
    def __init__(self, name, continent, capital, population,
                 area, coastline, currency, difficulty):
        self.name = name
        self.continent = continent
        self.capital = capital
        self.population = population
        self.population_margin = 0.3
        self.area = area
        self.area_margin = 0.5
        self.coastline = coastline
        self.coastine_margin = 0.5
        self.currency = currency
        self.difficulty = difficulty

    def __repr__(self):
        txt = self.name + os.linesep
        txt = txt + "Continent: " + self.continent + os.linesep
        txt = txt + "Capital: " + self.capital + os.linesep
        txt = txt + "Population: " + str(int(self.population/1000000)) + "M" + os.linesep
        txt = txt + "Area: " + str(self.area) + os.linesep
        txt = txt + "Coastline: " + str(self.coastline) + os.linesep
        txt = txt + "Currency: " + self.currency
        return txt
      
    def CheckCapital(self, capital):
        return stringsMatch(self.capital, capital)

    def CheckContinent(self, continent):
        return stringsMatch(self.continent, continent)

    def CheckCurrency(self, currency):
        return stringsMatch(self.currency, currency)

    def CheckPopulation(self, population):
        return isclose(self.population, population, self.population_margin)
            
    def CheckArea(self, area):
        return isclose(self.area, area, self.area_margin)       

    def CheckCoastline(self, coastline):
        if (self.coastline == 0 and coastline == 0) or (self.coastline != 0 and coastline != 0):
            return True
        else:
            return False
        
    def Quiz(self, player):
        answer = input("In which continent is the country of " + self.name + " ? ")
        if self.CheckContinent(answer):
            print ("This is correct")
            player.points += 1
            player.knowledge[self.name][0] = 1
        else:
            print ("This is wrong")
            player.knowledge[self.name][0] = 0
        print (" The answer was: ", self.continent)
        
        answer = input("What is the capital of " + self.name + " ? ")
        if self.CheckCapital(answer):
            print ("This is correct")
            player.points += 1
            player.knowledge[self.name][1] = 1
        else:
            print ("This is wrong")
            player.knowledge[self.name][1] = 0
        print (" The answer was: ", self.capital)

        answer = int(input("What is the population of " + self.name + " (in millions ?) ")) * 1000000
        if self.CheckPopulation(answer):
            print ("This is correct")
            player.points += 1
            player.knowledge[self.name][2] = 1
        else:
            print ("This is wrong")
            player.knowledge[self.name][2] = 0
        print (" The answer was: ", self.population)

        answer = int(input("What is the area of " + self.name + " ? "))
        if self.CheckArea(answer):
            print ("This is correct")
            player.points += 1
            player.knowledge[self.name][3] = 1
        else:
            print ("This is wrong")
            player.knowledge[self.name][3] = 0
        print (" The answer was: ", self.area)       

        answer = int(input("What is the coastline of " + self.name + " ? "))
        if self.CheckCoastline(answer):
            print ("This is correct")
            player.points += 1
            player.knowledge[self.name][4] = 1
        else:
            print ("This is wrong")
            player.knowledge[self.name][4] = 0
        print (" The answer was: ", self.coastline)

        answer = input("What is the currency of " + self.name + " ? ")
        if self.CheckCurrency(answer):
            print ("This is correct")
            player.points += 1
            player.knowledge[self.name][5] = 1
        else:
            print ("This is wrong")
            player.knowledge[self.name][5] = 0
        print (" The answer was: ", self.currency)       
    
# -------------------------------------------------
# Initialize global data, define useful functions
# -------------------------------------------------

Countries = {}
Players = []
Nb_Questions = 10

def CountryLoader():
    for row in range(1, worksheet.nrows):
        country = worksheet.row_values(row)         
        Countries[str(country[0])] = Country( 
            country[0],         # name
            country[1],         # continent
            country[2],         # capital
            country[3],         # population
            country[4],         # area
            country[5],         # coastline
            country[6],         # currency
            country[8])         # difficulty
     
# -------------------------------------------------
# Load data from Excel sheet
# -------------------------------------------------
print ("Loading country data...")
CountryLoader()
print (str(worksheet.nrows) + " countries loaded.")

# -------------------------------------------------
# Print all country data to test
# -------------------------------------------------
##for country in Countries:
##    print (Countries.get(country))
##    print ("")

# -------------------------------------------------
# Select and initialize player
# -------------------------------------------------
name = str(input("Enter player name: "))
player = Player(name)

# -------------------------------------------------
# Start game loop
# -------------------------------------------------

for i in range(1, Nb_Questions):
    country = random.choice(list(Countries.values()))
    country.Quiz(player)

player.SaveData()



    
